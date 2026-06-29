"""Core logic for UA Offer Generator."""

from __future__ import annotations

import io
import re
from datetime import datetime
from typing import Sequence

import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# "Artikl" is the UA style + colour key, e.g. 6009833-001.
OUTPUT_COLUMNS = [
    "Artikl",
    "Size",
    "Název",
    "Local warehouse 101",
    "Local warehouse 501",
    "Central warehouse",
    "Celkem ks styl/barva",
    "Plný sizerun",
    "Top style",
    "Materiál",
    "ORDER",
    "MOC CZK",
    "MOC EUR",
    "EAN",
    "Gender",
    "Silhouette",
    "Fit",
    "End use",
    "Season",
    "C/O",
]

EXTRA_COLUMNS = [
    "Dostupnost ve vybraných skladech",
    "Total available",
    "Dostupné velikosti styl/barva",
    "Style code",
    "Color group",
    "Color name",
    "Color code",
    "Detail silhouette",
    "Composition",
]


# Internal master-data field names remain stable for compatibility with the
# current source files. Every user-facing table and export is translated by
# this mapping at its presentation boundary.
DISPLAY_COLUMN_LABELS = {
    "Artikl": "Style / colour",
    "Název": "Product name",
    "Materiál": "Material",
    "Celkem ks styl/barva": "Total units / style-colour",
    "Plný sizerun": "Full size run",
    "MOC CZK": "RRP CZK",
    "MOC EUR": "RRP EUR",
    "Sleva %": "Discount %",
    "Dostupnost ve vybraných skladech": "Selected warehouse availability",
    "Total available": "Total availability (all warehouses)",
    "Dostupné velikosti styl/barva": "Available sizes / style-colour",
    "Style code": "Style code",
    "Importovaný požadavek": "Imported request",
    "Nedostupný artikl": "Unavailable style / colour",
    "Nedostupný název": "Unavailable product",
    "Alternativa – artikl": "Alternative style / colour",
    "Alternativa – název": "Alternative product",
    "Pořadí alternativy": "Alternative rank",
    "Typ alternativy": "Alternative type",
    "Úroveň shody": "Match level",
    "Skóre shody": "Match score",
    "Důvod doporučení": "Recommendation reason",
    "MOC nedostupného CZK": "Original RRP CZK",
    "MOC nedostupného EUR": "Original RRP EUR",
    "Barva nedostupného": "Original color group",
    "Celkem ks všechny sklady": "Total units (all warehouses)",
    "Dostupné velikosti ve vybraných skladech": "Available sizes in selected warehouses",
    "Velikosti v masteru": "Sizes in master",
    "Počet EAN": "EAN count",
    "Stav": "Status",
    "Poznámka": "Note",
}


def to_english_display(frame: pd.DataFrame) -> pd.DataFrame:
    """Return a copy with English user-facing headers, preserving row data."""
    return frame.rename(columns=DISPLAY_COLUMN_LABELS).copy()

SIZE_ORDER = {
    "XXS": 10,
    "XS": 20,
    "SM": 30,
    "S": 30,
    "MD": 40,
    "M": 40,
    "LG": 50,
    "L": 50,
    "XL": 60,
    "XXL": 70,
    "2XL": 70,
    "XXXL": 80,
    "3XL": 80,
    "4XL": 90,
    "5XL": 100,
    "OSFA": 900,
    "OSFM": 900,
}

TRUE_MARKERS = {"1", "true", "yes", "y", "ano", "x", "top", "top style"}
CORE_SIZE_RUNS = {
    "mens": {"S", "M", "L", "XL", "2XL"},
    "womens": {"XS", "S", "M", "L", "XL"},
}


def normalize_ean(value) -> str:
    """Return a safe text EAN without decimals or spaces."""
    if pd.isna(value):
        return ""
    text = str(value).strip()
    if not text or text.casefold() in {"nan", "none"}:
        return ""
    if re.fullmatch(r"\d+\.0", text):
        text = text[:-2]
    digits = re.sub(r"\D", "", text)
    return digits if len(digits) >= 8 else text


def read_excel(path_or_buffer) -> pd.DataFrame:
    """Read the first worksheet into a DataFrame."""
    from openpyxl import load_workbook

    wb = load_workbook(path_or_buffer, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = ws.iter_rows(values_only=True)
    try:
        header_row = next(rows)
    except StopIteration:
        return pd.DataFrame()

    headers: list[str] = []
    seen: dict[str, int] = {}
    for idx, value in enumerate(header_row, start=1):
        name = str(value).strip() if value is not None else f"Column {idx}"
        if name in seen:
            seen[name] += 1
            name = f"{name}_{seen[name]}"
        else:
            seen[name] = 1
        headers.append(name)

    return pd.DataFrame(list(rows), columns=headers)


STYLE_SELECTOR_RE = re.compile(r"^(\d{5,10})(?:-([A-Z0-9]{2,6}))?$")


def normalize_style_selector(value) -> str:
    """Normalize an imported UA style reference.

    Accepted values are either a base style (for example ``1326799``) or an
    exact style/colour reference (for example ``1326799-036``). Values that
    do not follow this format are ignored; this also makes a header such as
    ``Artikl`` harmless.
    """
    if value is None or pd.isna(value):
        return ""

    text = str(value).strip().upper()
    if re.fullmatch(r"\d+\.0", text):
        text = text[:-2]
    text = (
        text.replace("–", "-")
        .replace("—", "-")
        .replace("−", "-")
        .replace("/", "-")
    )
    text = re.sub(r"\s+", "", text)
    match = STYLE_SELECTOR_RE.fullmatch(text)
    if not match:
        return ""

    style_code, colour_code = match.groups()
    return f"{style_code}-{colour_code}" if colour_code else style_code


def normalize_style_selectors(values: Sequence[object]) -> list[str]:
    """Return unique valid imported style references in source order."""
    result: list[str] = []
    seen: set[str] = set()
    for value in values:
        selector = normalize_style_selector(value)
        if selector and selector not in seen:
            result.append(selector)
            seen.add(selector)
    return result


def read_style_selectors_excel(path_or_buffer) -> list[str]:
    """Read style references from every populated cell of the first XLSX sheet.

    The import therefore works for the requested one-column layout with values
    below each other, whether the file contains a column heading or not.
    """
    from openpyxl import load_workbook

    wb = load_workbook(path_or_buffer, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    values = (value for row in ws.iter_rows(values_only=True) for value in row)
    return normalize_style_selectors(values)


def style_selection_mask(data: pd.DataFrame, style_selectors: Sequence[object]) -> pd.Series:
    """Match exact style/colour references or all colours of a base style.

    An imported base style has precedence over the generic colour filter: it
    intentionally includes every available colour belonging to that base style.
    """
    selectors = normalize_style_selectors(style_selectors)
    if not selectors:
        return pd.Series(True, index=data.index)

    exact_articles = {value for value in selectors if "-" in value}
    base_styles = {value for value in selectors if "-" not in value}

    articles = (
        clean_text(data["Artikl"])
        .str.upper()
        .str.replace("–", "-", regex=False)
        .str.replace("—", "-", regex=False)
        .str.replace(" ", "", regex=False)
    )
    style_codes = clean_text(data["Style code"]).str.upper()
    return articles.isin(exact_articles) | style_codes.isin(base_styles)


def unmatched_style_selectors(data: pd.DataFrame, style_selectors: Sequence[object]) -> list[str]:
    """Return valid imported references that are absent from master data."""
    selectors = normalize_style_selectors(style_selectors)
    if not selectors:
        return []

    article_values = set(
        clean_text(data["Artikl"])
        .str.upper()
        .str.replace("–", "-", regex=False)
        .str.replace("—", "-", regex=False)
        .str.replace(" ", "", regex=False)
        .tolist()
    )
    style_values = set(clean_text(data["Style code"]).str.upper().tolist())
    return [
        value for value in selectors
        if (value not in article_values if "-" in value else value not in style_values)
    ]


NO_STOCK_IMPORT_COLUMNS = [
    "Importovaný požadavek",
    "Artikl",
    "Název",
    "Gender",
    "Division",
    "Segment",
    "Silhouette",
    "Detail silhouette",
    "Fit",
    "End use",
    "Season",
    "C/O",
    "Color group",
    "Color name",
    "Color code",
    "Materiál",
    "Composition",
    "Velikosti v masteru",
    "Počet EAN",
    "Local warehouse 101",
    "Local warehouse 501",
    "Central warehouse",
    "Celkem ks všechny sklady",
    "MOC CZK",
    "MOC EUR",
]


def _normalized_article_series(data: pd.DataFrame) -> pd.Series:
    """Return normalized UA style-colour keys for matching imported values."""
    return (
        clean_text(data["Artikl"])
        .str.upper()
        .str.replace("–", "-", regex=False)
        .str.replace("—", "-", regex=False)
        .str.replace("−", "-", regex=False)
        .str.replace("/", "-", regex=False)
        .str.replace(" ", "", regex=False)
    )


def _first_nonempty_value(group: pd.DataFrame, column: str):
    """Get the first meaningful value from a style-colour group."""
    if column not in group.columns:
        return ""
    for value in group[column].tolist():
        if pd.notna(value) and str(value).strip() not in {"", "nan", "None"}:
            return value
    return ""


def imported_no_stock_report(
    data: pd.DataFrame,
    style_selectors: Sequence[object],
) -> pd.DataFrame:
    """Build one row per imported style-colour with zero stock everywhere.

    The report intentionally ignores the currently selected warehouse filter.
    It checks the complete combined stock (101 + 501 + Central) so it answers
    whether an imported item is unavailable at *all* uploaded warehouses.

    Exact imported values (for example ``1326799-036``) inspect that one
    style-colour. A base style (``1326799``) expands to all its colours and
    reports every colour that has zero stock everywhere.
    """
    selectors = normalize_style_selectors(style_selectors)
    if not selectors or data.empty:
        return pd.DataFrame(columns=NO_STOCK_IMPORT_COLUMNS)

    article_keys = _normalized_article_series(data)
    style_codes = clean_text(data["Style code"]).str.upper()

    # Preserve the import order, while avoiding duplicate rows when the import
    # contains both a base style and an exact style-colour value.
    requested_by_article: dict[str, list[str]] = {}
    article_order: list[str] = []
    for selector in selectors:
        selector_mask = article_keys.eq(selector) if "-" in selector else style_codes.eq(selector)
        for article_key in article_keys.loc[selector_mask].drop_duplicates().tolist():
            if article_key not in requested_by_article:
                requested_by_article[article_key] = []
                article_order.append(article_key)
            requested_by_article[article_key].append(selector)

    rows: list[dict[str, object]] = []
    for article_key in article_order:
        group = data.loc[article_keys.eq(article_key)].copy()
        if group.empty:
            continue

        stock_101 = int(round(to_number(group["Local warehouse 101"]).sum()))
        stock_501 = int(round(to_number(group["Local warehouse 501"]).sum()))
        stock_central = int(round(to_number(group["Central warehouse"]).sum()))
        stock_total = stock_101 + stock_501 + stock_central
        if stock_total != 0:
            continue

        sizes = sorted(
            {value for value in clean_text(group["Size"]).tolist() if value},
            key=size_sort_value,
        )
        rows.append(
            {
                "Importovaný požadavek": ", ".join(requested_by_article[article_key]),
                "Artikl": _first_nonempty_value(group, "Artikl"),
                "Název": _first_nonempty_value(group, "Název"),
                "Gender": _first_nonempty_value(group, "Gender"),
                "Division": _first_nonempty_value(group, "Division"),
                "Segment": _first_nonempty_value(group, "Segment"),
                "Silhouette": _first_nonempty_value(group, "Silhouette"),
                "Detail silhouette": _first_nonempty_value(group, "Detail silhouette"),
                "Fit": _first_nonempty_value(group, "Fit"),
                "End use": _first_nonempty_value(group, "End use"),
                "Season": _first_nonempty_value(group, "Season"),
                "C/O": _first_nonempty_value(group, "C/O"),
                "Color group": _first_nonempty_value(group, "Color group"),
                "Color name": _first_nonempty_value(group, "Color name"),
                "Color code": _first_nonempty_value(group, "Color code"),
                "Materiál": _first_nonempty_value(group, "Materiál"),
                "Composition": _first_nonempty_value(group, "Composition"),
                "Velikosti v masteru": ", ".join(sizes),
                "Počet EAN": int(len(group)),
                "Local warehouse 101": stock_101,
                "Local warehouse 501": stock_501,
                "Central warehouse": stock_central,
                "Celkem ks všechny sklady": stock_total,
                "MOC CZK": float(to_number(group["MOC CZK"]).max()),
                "MOC EUR": float(to_number(group["MOC EUR"]).max()),
            }
        )

    return pd.DataFrame(rows, columns=NO_STOCK_IMPORT_COLUMNS)


def write_import_no_stock_excel(
    no_stock_df: pd.DataFrame,
    missing_master_selectors: Sequence[object] = (),
    title: str = "Imported UA products without stock",
) -> bytes:
    """Write a downloadable control report for imported styles.

    The first worksheet lists style-colours existing in master data but absent
    from every uploaded warehouse. A second sheet lists valid imported inputs
    that do not exist in master data at all.
    """
    wb = Workbook()
    header_fill = PatternFill("solid", fgColor="1F2937")
    header_font = Font(color="FFFFFF", bold=True)
    title_font = Font(size=14, bold=True, color="111827")
    thin_side = Side(style="thin", color="D1D5DB")
    border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)

    def write_sheet(
        sheet_name: str,
        report_title: str,
        note: str,
        frame: pd.DataFrame,
    ) -> None:
        ws = wb.active if wb.active.title == "Sheet" else wb.create_sheet()
        ws.title = sheet_name
        ws.cell(row=1, column=1, value=report_title).font = title_font
        ws.cell(row=2, column=1, value=f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}")
        ws.cell(row=3, column=1, value=note)

        for col_idx, column_name in enumerate(frame.columns, start=1):
            cell = ws.cell(row=4, column=col_idx, value=column_name)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = border

        integer_columns = {
            "EAN count",
            "Local warehouse 101",
            "Local warehouse 501",
            "Central warehouse",
            "Total units (all warehouses)",
        }
        money_columns = {"RRP CZK", "RRP EUR"}
        for row_idx, row in enumerate(frame.itertuples(index=False), start=5):
            for col_idx, value in enumerate(row, start=1):
                header = frame.columns[col_idx - 1]
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.border = border
                cell.alignment = Alignment(vertical="center", wrap_text=header in {"Product name", "Composition"})
                if header in integer_columns:
                    cell.number_format = "#,##0"
                elif header in money_columns:
                    cell.number_format = "#,##0.00"

        ws.freeze_panes = "A5"
        ws.auto_filter.ref = f"A4:{get_column_letter(len(frame.columns))}{4 + len(frame)}"
        widths = {
            "Imported request": 24,
            "Style / colour": 18,
            "Product name": 34,
            "Gender": 12,
            "Division": 16,
            "Segment": 16,
            "Silhouette": 16,
            "Detail silhouette": 20,
            "Fit": 16,
            "End use": 16,
            "Season": 12,
            "C/O": 10,
            "Color group": 16,
            "Color name": 22,
            "Color code": 12,
            "Material": 24,
            "Composition": 36,
            "Sizes in master": 24,
            "EAN count": 12,
            "Local warehouse 101": 18,
            "Local warehouse 501": 18,
            "Central warehouse": 18,
            "Total units (all warehouses)": 24,
            "RRP CZK": 13,
            "RRP EUR": 13,
            "Status": 20,
            "Note": 36,
        }
        for idx, column_name in enumerate(frame.columns, start=1):
            ws.column_dimensions[get_column_letter(idx)].width = widths.get(column_name, 16)
        ws.row_dimensions[1].height = 22
        ws.row_dimensions[4].height = 32

    report = no_stock_df.copy()
    if report.empty:
        report = pd.DataFrame(columns=NO_STOCK_IMPORT_COLUMNS)
    write_sheet(
        "Unavailable imports",
        title,
        "Products exist in master data but have zero total quantity across Local warehouse 101, Local warehouse 501 and Central warehouse.",
        to_english_display(report),
    )

    missing = normalize_style_selectors(missing_master_selectors)
    if missing:
        missing_df = pd.DataFrame(
            {
                "Imported request": missing,
                "Status": "Not found in master data",
                "Note": "No matching style or style-colour reference exists in the uploaded master data.",
            }
        )
        write_sheet(
            "Not in master data",
            "Imported references not found in master data",
            "These values could not be matched to the uploaded master data, so product attributes are not available.",
            missing_df,
        )

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()


SUBSTITUTION_TARGET_COLUMNS = [
    "Importovaný požadavek",
    "Artikl",
    "Název",
    "Gender",
    "Division",
    "Segment",
    "Silhouette",
    "Detail silhouette",
    "Fit",
    "End use",
    "Season",
    "C/O",
    "Color group",
    "Color name",
    "Color code",
    "Materiál",
    "Composition",
    "MOC CZK",
    "MOC EUR",
    "Celkem ks všechny sklady",
    "Dostupné velikosti ve vybraných skladech",
]

SUBSTITUTION_ALTERNATIVE_COLUMNS = [
    "Importovaný požadavek",
    "Nedostupný artikl",
    "Nedostupný název",
    "Alternativa – artikl",
    "Alternativa – název",
    "Pořadí alternativy",
    "Typ alternativy",
    "Úroveň shody",
    "Skóre shody",
    "Důvod doporučení",
    "MOC nedostupného CZK",
    "MOC nedostupného EUR",
    "Barva nedostupného",
    "Gender",
    "Division",
    "Segment",
    "Silhouette",
    "Detail silhouette",
    "Fit",
    "End use",
    "Materiál",
    "Color group",
    "Color name",
    "MOC CZK",
    "MOC EUR",
    "Dostupnost ve vybraných skladech",
    "Celkem ks všechny sklady",
    "Dostupné velikosti ve vybraných skladech",
    "Plný sizerun",
    "Top style",
]

ALTERNATIVE_ENGINE_VERSION = "v13"



def _comparison_text(value: object) -> str:
    """Normalize an attribute for substitution matching."""
    if value is None or pd.isna(value):
        return ""
    return str(value).strip().casefold()


def _base_style_key(value: object) -> str:
    """Return the normalized UA base-style key derived directly from Artikl.

    This intentionally does not depend on the helper ``Style code`` column.
    It is the hard safety check that prevents another colourway of the same
    base style from ever entering the alternatives list.
    """
    if value is None or pd.isna(value):
        return ""
    text = str(value).strip().upper()
    text = (
        text.replace("–", "-")
        .replace("—", "-")
        .replace("−", "-")
        .replace("/", "-")
    )
    text = re.sub(r"\s+", "", text)
    return text.split("-", 1)[0] if text else ""


def _base_style_series(values: pd.Series) -> pd.Series:
    """Vectorized counterpart to ``_base_style_key``."""
    return values.map(_base_style_key)


def _same_attribute(target: pd.Series, candidate: pd.Series, column: str) -> bool:
    target_value = _comparison_text(target.get(column, ""))
    candidate_value = _comparison_text(candidate.get(column, ""))
    return bool(target_value and candidate_value and target_value == candidate_value)


def _article_catalog(
    data: pd.DataFrame,
    selected_warehouses: Sequence[str],
) -> pd.DataFrame:
    """Build one product record per UA style / colour for substitution logic.

    The stock source for recommendations is deliberately the same set of
    warehouses currently selected in the offer filters. This prevents a
    recommendation that is available only in an excluded warehouse.
    """
    if data.empty:
        return pd.DataFrame()

    enriched = add_stock_metrics(data, selected_warehouses)
    records: list[dict[str, object]] = []
    fields = [
        "Artikl", "Style code", "Název", "Gender", "Division", "Segment",
        "Silhouette", "Detail silhouette", "Fit", "End use", "Season", "C/O",
        "Color group", "Color name", "Color code", "Materiál", "Composition",
        "Top style", "Plný sizerun",
    ]
    for article, group in enriched.groupby("Artikl", sort=False):
        available_sizes = sorted(
            {
                standard_size(value)
                for value in group.loc[
                    group["Dostupnost ve vybraných skladech"] > 0, "Size"
                ].tolist()
                if standard_size(value)
            },
            key=size_sort_value,
        )
        record = {
            field: _first_nonempty_value(group, field)
            for field in fields
        }
        record["Artikl"] = article
        record["_article_key"] = _normalized_article_series(group).iloc[0]
        record["_base_style_key"] = _base_style_key(article)
        record["MOC CZK"] = float(to_number(group["MOC CZK"]).max())
        record["MOC EUR"] = float(to_number(group["MOC EUR"]).max())
        record["Dostupnost ve vybraných skladech"] = int(
            to_number(group["Dostupnost ve vybraných skladech"]).sum()
        )
        if "Total available" in group.columns:
            all_stock_total = to_number(group["Total available"]).sum()
        else:
            all_stock_cols = [
                column for column in ["Local warehouse 101", "Local warehouse 501", "Central warehouse"]
                if column in group.columns
            ]
            all_stock_total = group[all_stock_cols].apply(to_number).sum(axis=1).sum() if all_stock_cols else 0
        record["Celkem ks všechny sklady"] = int(all_stock_total)
        record["Dostupné velikosti ve vybraných skladech"] = ", ".join(available_sizes)
        records.append(record)

    return pd.DataFrame(records)


def _price_match_score(target_price: float, candidate_price: float) -> tuple[int, str | None]:
    """Score price proximity without making it a hard exclusion."""
    if target_price <= 0 or candidate_price <= 0:
        return 0, None
    difference = abs(candidate_price - target_price) / target_price
    if difference <= 0.05:
        return 15, "very close RRP"
    if difference <= 0.15:
        return 11, "similar RRP"
    if difference <= 0.25:
        return 7, "comparable RRP"
    if difference <= 0.50:
        return 3, "approximate RRP"
    return 0, None


PRODUCT_FAMILY_PATTERNS = (
    ("polo", re.compile(r"\bpolo\b", re.IGNORECASE)),
    ("t-shirt", re.compile(r"\btee\b|t[ -]?shirt", re.IGNORECASE)),
    ("shorts", re.compile(r"\bshorts?\b", re.IGNORECASE)),
    ("leggings", re.compile(r"\bleggings?\b", re.IGNORECASE)),
    ("pants", re.compile(r"\bpants?\b|\btrousers?\b", re.IGNORECASE)),
    ("jacket", re.compile(r"\bjacket\b", re.IGNORECASE)),
    ("hoodie", re.compile(r"\bhoodie\b|\bhooded\b", re.IGNORECASE)),
    ("vest", re.compile(r"\bvest\b", re.IGNORECASE)),
    ("bra", re.compile(r"\bbra\b", re.IGNORECASE)),
    ("skirt", re.compile(r"\bskirt\b", re.IGNORECASE)),
    ("visor", re.compile(r"\bvisor\b", re.IGNORECASE)),
    ("cap", re.compile(r"\bcap\b|\bhat\b", re.IGNORECASE)),
    ("socks", re.compile(r"\bsocks?\b", re.IGNORECASE)),
    ("shoes", re.compile(r"\bshoe\b|\bshoes\b|\bsneaker\b|\bcleat\b", re.IGNORECASE)),
)


def _product_family_tokens(value: object) -> set[str]:
    """Extract a small, practical product-family signal from a UA item name.

    Master data does not always expose a dedicated product-family field. This
    signal helps distinguish, for example, a golf polo from another short
    sleeve golf top. It is used as a strong ranking factor, never as the only
    matching criterion.
    """
    name = _comparison_text(value)
    if not name:
        return set()
    return {label for label, pattern in PRODUCT_FAMILY_PATTERNS if pattern.search(name)}


def _candidate_pool(
    candidates: pd.DataFrame,
    mask: pd.Series,
    label: str,
) -> tuple[pd.DataFrame, str] | None:
    """Return the first non-empty candidate pool with a human-readable label."""
    pool = candidates.loc[mask].copy()
    return (pool, label) if not pool.empty else None


def _match_new_style_pool(
    target: pd.Series,
    candidates: pd.DataFrame,
) -> tuple[pd.DataFrame, str] | None:
    """Find a practical *different-style* pool, preserving product identity.

    This path is used when the customer wants to retain the colour first and
    accepts a different UA base style. A polo therefore competes with other
    available polos before broader "tops" are considered.
    """
    if candidates.empty:
        return None

    target_family = _product_family_tokens(target.get("Název", ""))
    working = candidates.copy()
    family_used = False
    if target_family:
        family_mask = working["Název"].map(
            lambda value: bool(target_family & _product_family_tokens(value))
        )
        if family_mask.any():
            working = working.loc[family_mask].copy()
            family_used = True
    if working.empty:
        return None

    def matches(column: str, target_value: object, default: bool = False) -> pd.Series:
        target_text = _comparison_text(target_value)
        if not target_text:
            return pd.Series(default, index=working.index)
        return working[column].map(_comparison_text).eq(target_text)

    same_gender = matches("Gender", target.get("Gender", ""), default=True)
    same_division = matches("Division", target.get("Division", ""), default=True)
    compatible = same_gender & same_division
    if compatible.any():
        working = working.loc[compatible].copy()
    elif same_gender.any():
        working = working.loc[same_gender].copy()
    if working.empty:
        return None

    def current_matches(column: str, target_value: object) -> pd.Series:
        target_text = _comparison_text(target_value)
        if not target_text:
            return pd.Series(False, index=working.index)
        return working[column].map(_comparison_text).eq(target_text)

    same_end_use = current_matches("End use", target.get("End use", ""))
    same_silhouette = current_matches("Silhouette", target.get("Silhouette", ""))
    same_detail = current_matches("Detail silhouette", target.get("Detail silhouette", ""))
    same_color_group = current_matches("Color group", target.get("Color group", ""))
    category_match = same_detail | same_silhouette

    pools = [
        _candidate_pool(
            working,
            same_color_group & same_end_use & category_match,
            "different style + same colour + end use + silhouette",
        ),
        _candidate_pool(
            working,
            same_color_group & same_end_use,
            "different style + same colour + end use",
        ),
        _candidate_pool(
            working,
            same_color_group & category_match,
            "different style + same colour + silhouette",
        ),
        _candidate_pool(
            working,
            same_color_group,
            "different style + same colour",
        ),
        _candidate_pool(
            working,
            same_end_use & category_match,
            "different style + end use + silhouette (different colour)",
        ),
        _candidate_pool(
            working,
            same_end_use,
            "different style + end use (different colour)",
        ),
        _candidate_pool(
            working,
            category_match,
            "different style + silhouette (different colour)",
        ),
        _candidate_pool(
            working,
            pd.Series(True, index=working.index),
            "different style – broader alternative",
        ),
    ]
    return next((pool for pool in pools if pool is not None), None)


def _rank_alternative_candidates(
    target: pd.Series,
    candidates: pd.DataFrame,
    alternative_type: str,
    match_level: str,
) -> list[dict[str, object]]:
    """Rank one already-selected alternative path.

    ``alternative_type`` makes the two legitimate commercial choices visible:
    retain the same style and change colour, or retain colour and change style.
    """
    ranked: list[dict[str, object]] = []
    target_family = _product_family_tokens(target.get("Název", ""))

    for _, candidate in candidates.iterrows():
        score = 0
        reasons: list[str] = []
        candidate_family = _product_family_tokens(candidate.get("Název", ""))
        matching_family = sorted(target_family & candidate_family)

        if alternative_type == "Same style, different colour":
            score += 140
            reasons.append("same style, another available colour")
        else:
            reasons.append("different available style")

        if matching_family:
            score += 60
            reasons.append(f"same product family: {', '.join(matching_family)}")

        if _same_attribute(target, candidate, "Gender"):
            score += 35
            reasons.append("same gender")
        if _same_attribute(target, candidate, "Division"):
            score += 25
        if _same_attribute(target, candidate, "Segment"):
            score += 10
            reasons.append("same segment")
        if _same_attribute(target, candidate, "Silhouette"):
            score += 24
            reasons.append("same silhouette")
        if _same_attribute(target, candidate, "Detail silhouette"):
            score += 30
            reasons.append("same detail silhouette")
        if _same_attribute(target, candidate, "Fit"):
            score += 12
            reasons.append("same fit")
        if _same_attribute(target, candidate, "End use"):
            score += 28
            reasons.append("same end use")
        if _same_attribute(target, candidate, "Materiál"):
            score += 7
            reasons.append("same material")
        if _same_attribute(target, candidate, "Color group"):
            score += 35
            reasons.append("same color group")
        if _same_attribute(target, candidate, "Color name"):
            score += 12
            reasons.append("same color name")
        if _same_attribute(target, candidate, "Season"):
            score += 3

        price_score, price_reason = _price_match_score(
            float(target.get("MOC CZK", 0) or 0),
            float(candidate.get("MOC CZK", 0) or 0),
        )
        score += price_score
        if price_reason:
            reasons.append(price_reason)
        if bool(candidate.get("Top style", False)):
            score += 3
            reasons.append("Top style")

        ranked.append(
            {
                "_score": score,
                "_selected_qty": int(candidate.get("Dostupnost ve vybraných skladech", 0) or 0),
                "_candidate": candidate,
                "_reasons": reasons,
                "_alternative_type": alternative_type,
                "_match_level": match_level,
            }
        )

    ranked.sort(
        key=lambda item: (
            -item["_score"],
            -item["_selected_qty"],
            str(item["_candidate"].get("Artikl", "")),
        )
    )
    return ranked


def _alternative_rows(
    target: pd.Series,
    ranked: Sequence[dict[str, object]],
    start_order: int = 1,
) -> list[dict[str, object]]:
    """Convert ranked candidates into the display/export structure."""
    rows: list[dict[str, object]] = []
    for order, item in enumerate(ranked, start=start_order):
        candidate = item["_candidate"]
        reasons = item["_reasons"] or ["closest available alternative based on matching attributes"]
        rows.append(
            {
                "Importovaný požadavek": target.get("Importovaný požadavek", ""),
                "Nedostupný artikl": target.get("Artikl", ""),
                "Nedostupný název": target.get("Název", ""),
                "Alternativa – artikl": candidate.get("Artikl", ""),
                "Alternativa – název": candidate.get("Název", ""),
                "Pořadí alternativy": order,
                "Typ alternativy": item["_alternative_type"],
                "Úroveň shody": item["_match_level"],
                "Skóre shody": item["_score"],
                "Důvod doporučení": "; ".join(reasons),
                "MOC nedostupného CZK": target.get("MOC CZK", 0),
                "MOC nedostupného EUR": target.get("MOC EUR", 0),
                "Barva nedostupného": target.get("Color group", ""),
                "Gender": candidate.get("Gender", ""),
                "Division": candidate.get("Division", ""),
                "Segment": candidate.get("Segment", ""),
                "Silhouette": candidate.get("Silhouette", ""),
                "Detail silhouette": candidate.get("Detail silhouette", ""),
                "Fit": candidate.get("Fit", ""),
                "End use": candidate.get("End use", ""),
                "Materiál": candidate.get("Materiál", ""),
                "Color group": candidate.get("Color group", ""),
                "Color name": candidate.get("Color name", ""),
                "MOC CZK": candidate.get("MOC CZK", 0),
                "MOC EUR": candidate.get("MOC EUR", 0),
                "Dostupnost ve vybraných skladech": candidate.get("Dostupnost ve vybraných skladech", 0),
                "Celkem ks všechny sklady": candidate.get("Celkem ks všechny sklady", 0),
                "Dostupné velikosti ve vybraných skladech": candidate.get("Dostupné velikosti ve vybraných skladech", ""),
                "Plný sizerun": candidate.get("Plný sizerun", ""),
                "Top style": "Yes" if bool(candidate.get("Top style", False)) else "",
            }
        )
    return rows


def _recommendations_for_target(
    target: pd.Series,
    catalog: pd.DataFrame,
    max_alternatives: int,
    alternative_strategy: str = "both",
) -> pd.DataFrame:
    """Return commercially useful available alternatives for one unavailable item.

    The unavailable *exact* style-colour is always excluded. Two valid routes
    are supported:

    * ``same_style``: retain the base style and offer another available colour;
    * ``same_color_new_style``: retain the colour first and find another style;
    * ``both``: show both routes in one short, balanced list.

    This reflects how a customer normally decides: keep fit/price/style and
    accept a colour change, or keep the colour and accept a different product.
    """
    if catalog.empty:
        return pd.DataFrame(columns=SUBSTITUTION_ALTERNATIVE_COLUMNS)

    strategy = str(alternative_strategy or "both").strip().casefold()
    if strategy not in {"both", "same_style", "same_color_new_style"}:
        strategy = "both"

    target_article_key = _comparison_text(target.get("_article_key", ""))
    target_base_style = _base_style_key(target.get("Artikl", "")) or _base_style_key(target.get("Style code", ""))

    available = catalog.loc[catalog["Dostupnost ve vybraných skladech"] > 0].copy()
    if available.empty:
        return pd.DataFrame(columns=SUBSTITUTION_ALTERNATIVE_COLUMNS)
    if "_base_style_key" not in available.columns:
        available["_base_style_key"] = _base_style_series(available["Artikl"])
    else:
        available["_base_style_key"] = available["_base_style_key"].map(_base_style_key)

    # This is the non-negotiable guard: never return the exact unavailable
    # style/colour. Other colours of that same style remain legitimate when
    # the user wants to preserve the cut, price and product identity.
    if target_article_key:
        available = available.loc[
            available["_article_key"].map(_comparison_text).ne(target_article_key)
        ].copy()
    if available.empty:
        return pd.DataFrame(columns=SUBSTITUTION_ALTERNATIVE_COLUMNS)

    same_style_ranked: list[dict[str, object]] = []
    new_style_ranked: list[dict[str, object]] = []

    if target_base_style:
        same_style_candidates = available.loc[
            available["_base_style_key"].eq(target_base_style)
        ].copy()
        if not same_style_candidates.empty:
            same_style_ranked = _rank_alternative_candidates(
                target,
                same_style_candidates,
                "Same style, different colour",
                "same style, another available colour",
            )

        new_style_candidates = available.loc[
            available["_base_style_key"].ne(target_base_style)
        ].copy()
    else:
        new_style_candidates = available

    selected_new_style = _match_new_style_pool(target, new_style_candidates)
    if selected_new_style is not None:
        new_style_pool, new_style_label = selected_new_style
        new_style_ranked = _rank_alternative_candidates(
            target,
            new_style_pool,
            "Different style, same colour",
            new_style_label,
        )

    limit = max(1, int(max_alternatives))
    chosen: list[dict[str, object]] = []
    if strategy == "same_style":
        chosen = same_style_ranked[:limit]
    elif strategy == "same_color_new_style":
        chosen = new_style_ranked[:limit]
    elif same_style_ranked and new_style_ranked:
        # Keep both commercial routes visible. For 5 suggestions this means
        # 2 colourways of the same style and 3 different styles in the same
        # colour; use the other route to fill any empty allocation.
        same_limit = max(1, limit // 2)
        new_limit = limit - same_limit
        chosen = same_style_ranked[:same_limit] + new_style_ranked[:new_limit]
        if len(chosen) < limit:
            used_same = min(len(same_style_ranked), same_limit)
            used_new = min(len(new_style_ranked), new_limit)
            remaining = same_style_ranked[used_same:] + new_style_ranked[used_new:]
            chosen.extend(remaining[: limit - len(chosen)])
    else:
        chosen = (same_style_ranked or new_style_ranked)[:limit]

    return pd.DataFrame(
        _alternative_rows(target, chosen),
        columns=SUBSTITUTION_ALTERNATIVE_COLUMNS,
    )

def imported_unavailable_with_alternatives(
    data: pd.DataFrame,
    style_selectors: Sequence[object],
    selected_warehouses: Sequence[str],
    max_alternatives: int = 5,
    alternative_strategy: str = "both",
) -> tuple[pd.DataFrame, pd.DataFrame]:
    """Create an unavailable-import list and ranked available alternatives.

    A requested item enters the first table only if it has zero stock across
    *all* uploaded warehouses. Recommendations then use the availability
    sources selected in the current offer criteria, so they can immediately be
    added to the offer.
    """
    unavailable = imported_no_stock_report(data, style_selectors).copy()
    if unavailable.empty:
        return (
            pd.DataFrame(columns=SUBSTITUTION_TARGET_COLUMNS),
            pd.DataFrame(columns=SUBSTITUTION_ALTERNATIVE_COLUMNS),
        )

    catalog = _article_catalog(data, selected_warehouses)
    if catalog.empty:
        return unavailable.reindex(columns=SUBSTITUTION_TARGET_COLUMNS), pd.DataFrame(columns=SUBSTITUTION_ALTERNATIVE_COLUMNS)

    target_rows: list[dict[str, object]] = []
    all_alternatives: list[pd.DataFrame] = []
    for _, unavailable_row in unavailable.iterrows():
        article_key = _normalized_article_series(
            pd.DataFrame({"Artikl": [unavailable_row.get("Artikl", "")]})
        ).iloc[0]
        target_candidates = catalog.loc[catalog["_article_key"].eq(article_key)]
        if target_candidates.empty:
            continue
        target_catalog = target_candidates.iloc[0].copy()
        target_catalog["Importovaný požadavek"] = unavailable_row.get("Importovaný požadavek", "")
        # Ensure the unavailable record retains the richer master metadata and
        # the full all-warehouse stock check from the no-stock report.
        target_row = {
            column: unavailable_row.get(column, target_catalog.get(column, ""))
            for column in SUBSTITUTION_TARGET_COLUMNS
        }
        target_rows.append(target_row)
        all_alternatives.append(
            _recommendations_for_target(
                target_catalog,
                catalog,
                max_alternatives,
                alternative_strategy=alternative_strategy,
            )
        )

    targets = pd.DataFrame(target_rows, columns=SUBSTITUTION_TARGET_COLUMNS)
    alternative_frames = [frame for frame in all_alternatives if not frame.empty]
    alternatives = (
        pd.concat(alternative_frames, ignore_index=True)
        if alternative_frames
        else pd.DataFrame(columns=SUBSTITUTION_ALTERNATIVE_COLUMNS)
    )
    return targets, alternatives


def add_selected_alternatives_to_offer(
    data: pd.DataFrame,
    base_offer: pd.DataFrame,
    selected_warehouses: Sequence[str],
    selected_alternatives: pd.DataFrame | None,
    replace_selected_unavailable: bool = True,
) -> pd.DataFrame:
    """Append user-chosen alternatives to the final offer.

    The normal import mode is deliberately preserved: base rows come from the
    imported file and bypass all standard filters. This helper then adds the
    explicit replacement choices from the alternatives table, also bypassing
    those regular filters, so the final export matches the commercial choice
    made in the Streamlit UI.

    When ``replace_selected_unavailable`` is true, only the exact unavailable
    style/colour that belongs to a checked alternative is removed. Other zero
    stock imports remain visible until the user selects a replacement for them.
    """
    if selected_alternatives is None or selected_alternatives.empty:
        return base_offer.copy()
    if "Alternativa – artikl" not in selected_alternatives.columns:
        return base_offer.copy()

    choices = selected_alternatives.copy()
    if "Přidat do nabídky" in choices.columns:
        choices = choices.loc[choices["Přidat do nabídky"].fillna(False).astype(bool)].copy()
    if choices.empty:
        return base_offer.copy()

    alternative_keys = set(
        _normalized_article_series(
            pd.DataFrame({"Artikl": choices["Alternativa – artikl"]})
        ).tolist()
    )
    alternative_keys.discard("")
    if not alternative_keys:
        return base_offer.copy()

    enriched = add_stock_metrics(data, selected_warehouses)
    enriched_keys = _normalized_article_series(enriched)
    selected_rows = enriched.loc[enriched_keys.isin(alternative_keys)].copy()
    if selected_rows.empty:
        return base_offer.copy()

    final_rows = base_offer.copy()
    if replace_selected_unavailable and "Nedostupný artikl" in choices.columns:
        unavailable_keys = set(
            _normalized_article_series(
                pd.DataFrame({"Artikl": choices["Nedostupný artikl"]})
            ).tolist()
        )
        unavailable_keys.discard("")
        if unavailable_keys and not final_rows.empty:
            base_keys = _normalized_article_series(final_rows)
            final_rows = final_rows.loc[~base_keys.isin(unavailable_keys)].copy()

    combined = pd.concat([final_rows, selected_rows], ignore_index=True, sort=False)
    if combined.empty:
        return combined

    combined["_article_sort"] = _normalized_article_series(combined)
    combined["_size_sort"] = combined["Size"].map(size_sort_value)
    dedupe_columns = [column for column in ["Artikl", "EAN"] if column in combined.columns]
    if dedupe_columns:
        combined = combined.drop_duplicates(subset=dedupe_columns, keep="first")
    combined = combined.sort_values(["_article_sort", "_size_sort", "Size", "EAN"])
    return combined.drop(
        columns=["_article_sort", "_size_sort", "_core_size", "_full_sizerun"],
        errors="ignore",
    )


def write_import_substitution_excel(
    unavailable_df: pd.DataFrame,
    alternatives_df: pd.DataFrame,
    missing_master_selectors: Sequence[object] = (),
    title: str = "Imported UA products unavailable and alternatives",
) -> bytes:
    """Write an English substitution workbook for imported products."""
    wb = Workbook()
    header_fill = PatternFill("solid", fgColor="1F2937")
    header_font = Font(color="FFFFFF", bold=True)
    title_font = Font(size=14, bold=True, color="111827")
    thin_side = Side(style="thin", color="D1D5DB")
    border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
    alternative_fill = PatternFill("solid", fgColor="E2F0D9")

    def write_sheet(sheet_name: str, report_title: str, note: str, frame: pd.DataFrame) -> None:
        ws = wb.active if wb.active.title == "Sheet" else wb.create_sheet()
        ws.title = sheet_name
        ws.cell(row=1, column=1, value=report_title).font = title_font
        ws.cell(row=2, column=1, value=f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}")
        ws.cell(row=3, column=1, value=note)
        for col_idx, column_name in enumerate(frame.columns, start=1):
            cell = ws.cell(row=4, column=col_idx, value=column_name)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = border

        integer_columns = {
            "Total units (all warehouses)", "Selected warehouse availability",
            "Alternative rank", "Match score", "EAN count",
        }
        money_columns = {"RRP CZK", "RRP EUR", "Original RRP CZK", "Original RRP EUR"}
        for row_idx, row in enumerate(frame.itertuples(index=False), start=5):
            for col_idx, value in enumerate(row, start=1):
                header = frame.columns[col_idx - 1]
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.border = border
                cell.alignment = Alignment(
                    vertical="center",
                    wrap_text=header in {"Product name", "Unavailable product", "Alternative product", "Recommendation reason", "Composition"},
                )
                if sheet_name == "Alternatives" and header.startswith("Alternative"):
                    cell.fill = alternative_fill
                if header in integer_columns:
                    cell.number_format = "#,##0"
                elif header in money_columns:
                    cell.number_format = "#,##0.00"

        ws.freeze_panes = "A5"
        ws.auto_filter.ref = f"A4:{get_column_letter(len(frame.columns))}{4 + len(frame)}"
        widths = {
            "Imported request": 24,
            "Style / colour": 18,
            "Unavailable style / colour": 20,
            "Alternative style / colour": 21,
            "Product name": 34,
            "Unavailable product": 34,
            "Alternative product": 34,
            "Recommendation reason": 44,
            "Gender": 12,
            "Division": 16,
            "Segment": 16,
            "Silhouette": 16,
            "Detail silhouette": 22,
            "Fit": 16,
            "End use": 16,
            "Season": 12,
            "C/O": 10,
            "Color group": 16,
            "Color name": 22,
            "Color code": 12,
            "Material": 24,
            "Composition": 34,
            "RRP CZK": 14,
            "RRP EUR": 14,
            "Original RRP CZK": 18,
            "Original RRP EUR": 18,
            "Original color group": 18,
            "Match level": 34,
            "Selected warehouse availability": 28,
            "Total units (all warehouses)": 25,
            "Available sizes in selected warehouses": 32,
            "Full size run": 14,
            "Top style": 12,
            "Alternative rank": 16,
            "Match score": 14,
            "Status": 22,
            "Note": 42,
        }
        for idx, column_name in enumerate(frame.columns, start=1):
            ws.column_dimensions[get_column_letter(idx)].width = widths.get(column_name, 16)
        ws.row_dimensions[1].height = 22
        ws.row_dimensions[4].height = 34

    unavailable = to_english_display(
        unavailable_df.copy().reindex(columns=SUBSTITUTION_TARGET_COLUMNS)
    )
    alternatives = to_english_display(
        alternatives_df.copy().reindex(columns=SUBSTITUTION_ALTERNATIVE_COLUMNS)
    )
    write_sheet(
        "Unavailable imports",
        title,
        "Imported style-colours that exist in master data but have zero stock in Local warehouse 101, Local warehouse 501 and Central warehouse.",
        unavailable,
    )
    write_sheet(
        "Alternatives",
        "Recommended available alternatives",
        "Alternatives can follow two routes: the same style in another available colour, or a different available style in the requested colour. The exact unavailable style-colour is always excluded. For a different style, product family (for example Polo), gender, division, colour, end use and silhouette are prioritised before fit, material, RRP and selected-warehouse availability.",
        alternatives,
    )

    missing = normalize_style_selectors(missing_master_selectors)
    if missing:
        missing_df = pd.DataFrame(
            {
                "Imported request": missing,
                "Status": "Not found in master data",
                "Note": "The product was not found in the current master data; attributes are unavailable, so relevant automatic alternatives cannot be created.",
            }
        )
        write_sheet(
            "Not in master data",
            "Imported references not found in master data",
            "These values could not be matched to the uploaded master data.",
            missing_df,
        )

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()

def find_col(df: pd.DataFrame, candidates: Sequence[str], required: bool = True) -> str | None:
    normalized = {str(c).strip().casefold(): c for c in df.columns}
    for candidate in candidates:
        key = candidate.strip().casefold()
        if key in normalized:
            return normalized[key]
    if required:
        raise KeyError(f"Missing required column. Tried: {', '.join(candidates)}")
    return None


def to_number(series: pd.Series) -> pd.Series:
    cleaned = (
        series.astype(str)
        .str.replace("\u00a0", "", regex=False)
        .str.replace(" ", "", regex=False)
        .str.replace(",", ".", regex=False)
        .replace({"nan": np.nan, "None": np.nan, "": np.nan})
    )
    return pd.to_numeric(cleaned, errors="coerce").fillna(0)


def clean_text(series: pd.Series) -> pd.Series:
    return series.fillna("").astype(str).str.strip()


def is_top_style(value) -> bool:
    return str(value).strip().casefold() in TRUE_MARKERS


def material_group(value) -> str:
    """Classify textile composition for quick material filters.

    Cotton blends are classified as Cotton. The performance group means a
    composition without cotton that contains common fibres such as polyester,
    elastane, nylon or polyamide.
    """
    text = str(value or "").casefold()
    if re.search(r"cotton|bavlna", text):
        return "Cotton"
    if re.search(r"polyester|elastane|spandex|nylon|polyamid|polyamide|polypropylene", text):
        return "Polyester / performance"
    return "Other / unclassified"


def standard_size(size: str) -> str:
    """Normalize UA size labels used in master data to standard size-run keys."""
    value = str(size or "").strip().upper().replace(" ", "")
    mapping = {
        "XXS": "XXS",
        "XS": "XS",
        "SM": "S",
        "S": "S",
        "MD": "M",
        "M": "M",
        "LG": "L",
        "L": "L",
        "XL": "XL",
        "XXL": "2XL",
        "2XL": "2XL",
        "XXXL": "3XL",
        "3XL": "3XL",
        "4XL": "4XL",
        "5XL": "5XL",
    }
    return mapping.get(value, value)


def first_available_cols_central(central_df: pd.DataFrame) -> list[str]:
    week_cols = [
        c for c in central_df.columns
        if re.match(r"^\s*week\s*\d+", str(c), flags=re.I)
    ]
    if len(week_cols) < 2:
        raise KeyError(
            "Central stock file must contain at least two Week columns, "
            "e.g. Week 25 and Week 26."
        )
    return week_cols[:2]


def aggregate_stock(
    df: pd.DataFrame,
    ean_candidates: Sequence[str],
    qty_candidates: Sequence[str],
    output_col: str,
) -> pd.DataFrame:
    ean_col = find_col(df, ean_candidates)
    qty_col = find_col(df, qty_candidates)
    tmp = pd.DataFrame(
        {
            "EAN": df[ean_col].map(normalize_ean),
            output_col: to_number(df[qty_col]),
        }
    )
    tmp = tmp[tmp["EAN"] != ""]
    return tmp.groupby("EAN", as_index=False)[output_col].sum()


def aggregate_central(central_df: pd.DataFrame) -> tuple[pd.DataFrame, list[str]]:
    ean_col = find_col(central_df, ["EAN", "EAN poslední"])
    week_cols = first_available_cols_central(central_df)

    tmp = pd.DataFrame({"EAN": central_df[ean_col].map(normalize_ean)})
    tmp["Central warehouse"] = sum(to_number(central_df[col]) for col in week_cols)
    tmp = tmp[tmp["EAN"] != ""]
    return tmp.groupby("EAN", as_index=False)["Central warehouse"].sum(), week_cols


def prepare_master(master_df: pd.DataFrame) -> pd.DataFrame:
    """Standardize master-data headers while accepting the current Czech export."""
    col_map = {
        "EAN": ["EAN"],
        "SKU": ["SKU"],
        "Artikl": ["Artikl", "Article", "ARTICLE_GENERIC"],
        "Size": ["Size US", "US Size", "Size", "Charakter.1", "Charakter"],
        "Název": ["Název", "Nazev", "Name", "ARTICLE_GENERIC_DESC"],
        # Master exports commonly use either MOC CZK / MOC EUR or CZK MOC / EUR MOC.
        # Keep both variants so retail prices are retained in preview and export.
        "MOC CZK": ["MOC CZK", "CZK MOC", "CZK RRP", "RRP CZK", "MOC", "Prodejní cena"],
        "MOC EUR": ["MOC EUR", "EUR MOC", "EUR RRP", "RRP EUR"],
        "Division": ["Division"],
        "Gender": ["Gender"],
        "Silhouette": ["Silhouette"],
        "Fit": ["Fit"],
        "Segment": ["Segment"],
        "End use": ["End use", "End Use"],
        "Detail silhouette": ["Detail silhouette", "Detail Silhouette"],
        "Season": ["Season"],
        "C/O": ["C/O", "CO"],
        "Color group": ["Color group", "Colour group", "Color Group", "Colour Group"],
        "Color name": ["Color name", "Colour name", "Color Name", "Colour Name"],
        "Color code": ["Color code", "Colour code", "Color Code", "Colour Code"],
        "Composition": ["Composition", "Materiál", "Material"],
        "Top style raw": ["Top style", "Top styles", "Top Style", "TOP STYLE"],
    }

    out = pd.DataFrame(index=master_df.index)
    required = {"EAN", "Artikl", "Size", "Název"}
    for canonical, candidates in col_map.items():
        col = find_col(master_df, candidates, required=canonical in required)
        out[canonical] = "" if col is None else master_df[col]

    out["EAN"] = out["EAN"].map(normalize_ean)
    text_cols = [col for col in out.columns if col not in {"MOC CZK", "MOC EUR"}]
    for col in text_cols:
        out[col] = clean_text(out[col])

    out["MOC CZK"] = to_number(out["MOC CZK"])
    out["MOC EUR"] = to_number(out["MOC EUR"])
    out["Style code"] = out["Artikl"].str.split("-", n=1).str[0].str.strip()
    out["Materiál"] = out["Composition"].map(material_group)

    # A marker in any EAN row labels the entire base style as a top style.
    # This lets the master-data editor mark one row instead of every size and colour.
    out["_top_style_flag"] = out["Top style raw"].map(is_top_style)
    style_flag = out.groupby("Style code")["_top_style_flag"].transform("max")
    out["Top style"] = style_flag.fillna(False).astype(bool)

    out = out[out["EAN"] != ""].drop_duplicates(subset=["EAN"], keep="first")
    return out.drop(columns=["Top style raw", "_top_style_flag"])


def build_dataset(
    master_df: pd.DataFrame,
    local101_df: pd.DataFrame,
    local501_df: pd.DataFrame,
    central_df: pd.DataFrame,
) -> tuple[pd.DataFrame, list[str]]:
    master = prepare_master(master_df)
    qty_candidates = [
        "Qty", "Quantity", "Množství", "Mnozstvi", "Počet", "Pocet",
        "Počet ks", "Pocet ks", "Stock", "Available", "Available Qty",
    ]
    stock101 = aggregate_stock(
        local101_df, ["EAN poslední", "EAN"], qty_candidates, "Local warehouse 101"
    )
    stock501 = aggregate_stock(
        local501_df, ["EAN poslední", "EAN"], qty_candidates, "Local warehouse 501"
    )
    central, central_cols = aggregate_central(central_df)

    data = master.merge(stock101, on="EAN", how="left")
    data = data.merge(stock501, on="EAN", how="left")
    data = data.merge(central, on="EAN", how="left")

    warehouse_cols = ["Local warehouse 101", "Local warehouse 501", "Central warehouse"]
    for col in warehouse_cols:
        data[col] = data[col].fillna(0).round(0).astype(int)
    data["Total available"] = data[warehouse_cols].sum(axis=1)
    return data, central_cols


def product_type_mask(
    data: pd.DataFrame,
    product_types: Sequence[str],
) -> pd.Series:
    if not product_types:
        return pd.Series(True, index=data.index)

    masks: list[pd.Series] = []
    division_apparel = data["Division"].str.casefold().eq("apparel")
    for product_type in product_types:
        if product_type == "Technical T-shirts":
            top_mask = data["Silhouette"].str.contains("tops", case=False, na=False)
            detail_mask = data["Detail silhouette"].str.contains(
                "sleeve|sleeveless|tee|t-shirt",
                case=False,
                na=False,
                regex=True,
            )
            mask = division_apparel & top_mask & detail_mask
            masks.append(mask)
        elif product_type == "Shorts":
            detail_mask = data["Detail silhouette"].str.contains("shorts", case=False, na=False)
            silhouette_mask = data["Silhouette"].str.contains("shorts", case=False, na=False)
            masks.append(division_apparel & (detail_mask | silhouette_mask))
        elif product_type == "Tops":
            masks.append(division_apparel & data["Silhouette"].str.contains("tops", case=False, na=False))
        elif product_type == "Footwear":
            masks.append(data["Division"].str.contains("footwear", case=False, na=False))
        elif product_type == "Accessories":
            masks.append(data["Division"].str.contains("accessories", case=False, na=False))

    if not masks:
        return pd.Series(True, index=data.index)

    result = masks[0].copy()
    for current in masks[1:]:
        result |= current
    return result


def color_mask(data: pd.DataFrame, selected_colors: Sequence[str]) -> pd.Series:
    if not selected_colors:
        return pd.Series(True, index=data.index)

    color_group = clean_text(data["Color group"])
    color_name = clean_text(data["Color name"])
    masks: list[pd.Series] = []

    for color in selected_colors:
        if color == "Black":
            masks.append(
                color_group.str.contains("black", case=False, na=False)
                | color_name.str.contains("black", case=False, na=False)
            )
        elif color == "Dark Blue":
            masks.append(
                color_name.str.contains("navy|midnight|academy", case=False, na=False, regex=True)
            )
        else:
            masks.append(color_group.str.contains(re.escape(color), case=False, na=False, regex=True))

    result = masks[0].copy()
    for current in masks[1:]:
        result |= current
    return result


def size_sort_value(size: str) -> int:
    key = str(size).strip().upper()
    if key in SIZE_ORDER:
        return SIZE_ORDER[key]
    try:
        return int(float(key) * 10)
    except Exception:
        return 1000


def add_stock_metrics(data: pd.DataFrame, selected_warehouses: Sequence[str]) -> pd.DataFrame:
    """Calculate the style/colour aggregate and full-size-run status.

    Every metric is calculated before the final row filters. Therefore the
    displayed style/colour stock is not reduced merely because another size is
    hidden by the minimum-per-EAN setting.
    """
    out = data.copy()
    warehouse_cols = ["Local warehouse 101", "Local warehouse 501", "Central warehouse"]
    selected = [col for col in selected_warehouses if col in warehouse_cols]
    if not selected:
        selected = warehouse_cols

    out["Dostupnost ve vybraných skladech"] = out[selected].sum(axis=1).round(0).astype(int)
    out["_core_size"] = out["Size"].map(standard_size)
    out["Celkem ks styl/barva"] = (
        out.groupby("Artikl")["Dostupnost ve vybraných skladech"].transform("sum").round(0).astype(int)
    )

    available_rows = out[out["Dostupnost ve vybraných skladech"] > 0].copy()
    size_counts = available_rows.groupby("Artikl")["_core_size"].nunique()
    out["Dostupné velikosti styl/barva"] = (
        out["Artikl"].map(size_counts).fillna(0).astype(int)
    )

    full_flags: dict[str, bool | None] = {}
    for artikel, group in out.groupby("Artikl", sort=False):
        genders = clean_text(group["Gender"]).str.casefold().unique().tolist()
        gender = genders[0] if genders else ""
        required = CORE_SIZE_RUNS.get(gender)
        if not required:
            full_flags[artikel] = None
            continue
        available = set(
            group.loc[group["Dostupnost ve vybraných skladech"] > 0, "_core_size"].tolist()
        )
        full_flags[artikel] = required.issubset(available)

    out["_full_sizerun"] = out["Artikl"].map(full_flags)
    out["Plný sizerun"] = out["_full_sizerun"].map(
        lambda value: "Yes" if value is True else ("No" if value is False else "")
    )
    return out


def apply_filters(
    data: pd.DataFrame,
    product_types: Sequence[str] = ("Technical T-shirts",),
    genders: Sequence[str] = ("Mens",),
    colors: Sequence[str] = ("Black", "Dark Blue"),
    price_min: float = 0,
    price_max: float = 999,
    price_eur_min: float = 0,
    price_eur_max: float = 999,
    selected_warehouses: Sequence[str] = ("Local warehouse 101", "Local warehouse 501", "Central warehouse"),
    min_total_available: int = 1,
    min_style_color_qty: int = 1,
    min_article_sizes: int = 1,
    seasons: Sequence[str] = (),
    end_uses: Sequence[str] = (),
    detail_silhouettes: Sequence[str] = (),
    fits: Sequence[str] = (),
    co_values: Sequence[str] = (),
    technical_material: bool = False,
    cotton_material: bool = False,
    only_top_styles: bool = False,
    only_full_sizerun: bool = False,
    selected_style_refs: Sequence[object] = (),
) -> pd.DataFrame:
    data = add_stock_metrics(data, selected_warehouses)

    # XLSX import is an explicit product-selection mode. Its contents take
    # precedence over every regular offer filter: product type, gender,
    # material, Top style, colour, MOC ranges, season, end use, silhouette,
    # fit, C/O and all stock / size-run thresholds. The selected warehouse
    # choice is still used to calculate the displayed availability, total
    # quantity per style-colour and sizerun status; it is not used to exclude
    # an imported product. This makes the imported file the single source of
    # truth for the resulting offer and also keeps zero-stock imports visible.
    if selected_style_refs:
        filtered = data.loc[style_selection_mask(data, selected_style_refs)].copy()
    else:
        mask = pd.Series(True, index=data.index)

        # Product type and material are independent filters. Material restrictions
        # are applied below only when one or both material checkboxes are selected.
        mask &= product_type_mask(data, product_types)

        if genders:
            mask &= data["Gender"].isin(genders)
        if colors:
            mask &= color_mask(data, colors)

        mask &= data["MOC CZK"].between(price_min, price_max, inclusive="both")
        mask &= data["MOC EUR"].between(price_eur_min, price_eur_max, inclusive="both")

        if seasons:
            mask &= data["Season"].isin(seasons)
        if end_uses:
            mask &= data["End use"].isin(end_uses)
        if detail_silhouettes:
            mask &= data["Detail silhouette"].isin(detail_silhouettes)
        if fits:
            mask &= data["Fit"].isin(fits)
        if co_values:
            mask &= data["C/O"].isin(co_values)

        if technical_material or cotton_material:
            material_options = set()
            if technical_material:
                material_options.add("Polyester / performance")
            if cotton_material:
                material_options.add("Cotton")
            mask &= data["Materiál"].isin(material_options)

        if only_top_styles:
            mask &= data["Top style"]

        if only_full_sizerun:
            mask &= data["_full_sizerun"].eq(True)

        mask &= data["Dostupnost ve vybraných skladech"] >= int(min_total_available)
        mask &= data["Celkem ks styl/barva"] >= int(min_style_color_qty)
        mask &= data["Dostupné velikosti styl/barva"] >= int(min_article_sizes)

        filtered = data[mask].copy()
    if filtered.empty:
        return filtered

    filtered["_size_sort"] = filtered["Size"].map(size_sort_value)
    filtered = (
        filtered.sort_values(["Artikl", "_size_sort", "Size", "EAN"])
        .drop(columns=["_size_sort", "_core_size", "_full_sizerun"], errors="ignore")
    )
    return filtered


def display_qty(value: int | float) -> int | str:
    try:
        number = int(round(float(value)))
    except Exception:
        return ""
    return "100+" if number > 100 else number


def _currency_flags(export_currency: str) -> tuple[bool, bool]:
    """Return whether CZK and EUR price columns should be included."""
    value = str(export_currency or "CZK + EUR").strip().upper()
    include_czk = value in {"CZK", "CZK + EUR", "BOTH", "ALL"}
    include_eur = value in {"EUR", "CZK + EUR", "BOTH", "ALL"}
    # Defensive fallback: keep both prices rather than silently exporting none.
    return (include_czk, include_eur) if include_czk or include_eur else (True, True)


def price_includes_vat(price_vat_mode: str) -> bool:
    """Return whether the requested final offer price is VAT-inclusive."""
    value = str(price_vat_mode or "Excl. VAT").strip().casefold()
    return value in {"incl. vat", "including vat", "with vat", "s dph", "včetně dph", "vcetne dph"}


def offer_price_column_name(currency: str, price_vat_mode: str) -> str:
    suffix = "incl. VAT" if price_includes_vat(price_vat_mode) else "excl. VAT"
    return f"Offer price {currency} {suffix}"


def to_offer_table(
    filtered: pd.DataFrame,
    include_extra_columns: bool = False,
    export_currency: str = "CZK + EUR",
    discount_percent: float | None = None,
    vat_rate: float = 0.21,
    price_vat_mode: str = "Excl. VAT",
) -> pd.DataFrame:
    """Build an English offer table and optional negotiated price columns.

    RRP values in the master are VAT-inclusive. The final export can show a
    discounted offer price either excluding VAT or including VAT. A ``None``
    discount keeps the preview neutral and shows only the RRP columns.
    """
    include_czk, include_eur = _currency_flags(export_currency)
    base = {
        "Style / colour": filtered["Artikl"],
        "Size": filtered["Size"],
        "Product name": filtered["Název"],
        "Local warehouse 101": filtered["Local warehouse 101"].map(display_qty),
        "Local warehouse 501": filtered["Local warehouse 501"].map(display_qty),
        "Central warehouse": filtered["Central warehouse"].map(display_qty),
        "Total units / style-colour": filtered["Celkem ks styl/barva"],
        "Full size run": filtered["Plný sizerun"],
        "Top style": filtered["Top style"].map(lambda value: "Yes" if bool(value) else ""),
        "Material": filtered["Materiál"],
        "ORDER": "",
    }

    if discount_percent is not None:
        discount_value = max(0.0, min(float(discount_percent), 100.0))
        vat_value = max(0.0, float(vat_rate))
        discounted_factor = 1 - discount_value / 100
        price_factor = discounted_factor if price_includes_vat(price_vat_mode) else discounted_factor / (1 + vat_value)
        base["Discount %"] = discount_value
        if include_czk:
            base["RRP CZK"] = filtered["MOC CZK"]
            base[offer_price_column_name("CZK", price_vat_mode)] = (
                filtered["MOC CZK"] * price_factor
            ).round(2)
        if include_eur:
            base["RRP EUR"] = filtered["MOC EUR"]
            base[offer_price_column_name("EUR", price_vat_mode)] = (
                filtered["MOC EUR"] * price_factor
            ).round(2)
    else:
        if include_czk:
            base["RRP CZK"] = filtered["MOC CZK"]
        if include_eur:
            base["RRP EUR"] = filtered["MOC EUR"]

    base.update(
        {
            "EAN": filtered["EAN"].astype(str),
            "Gender": filtered["Gender"],
            "Silhouette": filtered["Silhouette"],
            "Fit": filtered["Fit"],
            "End use": filtered["End use"],
            "Season": filtered["Season"],
            "C/O": filtered["C/O"],
        }
    )
    offer = pd.DataFrame(base)

    if include_extra_columns:
        for col in EXTRA_COLUMNS:
            if col in filtered.columns:
                offer[DISPLAY_COLUMN_LABELS.get(col, col)] = filtered[col]
    return offer

def write_offer_excel(
    offer_df: pd.DataFrame,
    title: str = "Under Armour Product Offer",
    discount_percent: float | None = None,
    export_currency: str = "CZK + EUR",
    vat_rate: float = 0.21,
    price_vat_mode: str = "Excl. VAT",
) -> bytes:
    """Write the customer-facing English offer workbook."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Offer"

    header_fill = PatternFill("solid", fgColor="1F2937")
    header_font = Font(color="FFFFFF", bold=True)
    order_fill = PatternFill("solid", fgColor="FFF2CC")
    offer_price_fill = PatternFill("solid", fgColor="E2F0D9")
    title_font = Font(size=14, bold=True, color="111827")
    thin_side = Side(style="thin", color="D1D5DB")
    border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)

    ws.cell(row=1, column=1, value=title).font = title_font
    ws.cell(row=2, column=1, value=f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}")
    if discount_percent is not None:
        ws.cell(
            row=3,
            column=1,
            value=(
                f"Currency: {export_currency} | Discount from VAT-inclusive RRP: "
                f"{float(discount_percent):g}% | VAT: {float(vat_rate) * 100:g}% | "
                f"Offer prices: {'including VAT' if price_includes_vat(price_vat_mode) else 'excluding VAT'}."
            ),
        )

    start_row = 4
    for col_idx, column_name in enumerate(offer_df.columns, start=1):
        cell = ws.cell(row=start_row, column=col_idx, value=column_name)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border

    integer_columns = {
        "Total units / style-colour",
        "Selected warehouse availability",
        "Total availability (all warehouses)",
        "Available sizes / style-colour",
    }
    money_columns = {"RRP CZK", "RRP EUR"}
    discounted_price_columns = {
        column for column in offer_df.columns if str(column).startswith("Offer price ")
    }
    money_columns |= discounted_price_columns
    for row_idx, row in enumerate(offer_df.itertuples(index=False), start=start_row + 1):
        for col_idx, value in enumerate(row, start=1):
            header = offer_df.columns[col_idx - 1]
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = border
            cell.alignment = Alignment(vertical="center")
            if header == "ORDER":
                cell.fill = order_fill
            elif header in discounted_price_columns:
                cell.fill = offer_price_fill
            if header == "EAN":
                cell.number_format = "@"
            elif header in money_columns:
                cell.number_format = "#,##0.00"
            elif header == "Discount %":
                cell.number_format = "0.0"
            elif header in integer_columns:
                cell.number_format = "#,##0"

    ws.freeze_panes = "A5"
    ws.auto_filter.ref = (
        f"A{start_row}:{get_column_letter(len(offer_df.columns))}{start_row + len(offer_df)}"
    )

    widths = {
        "Style / colour": 18,
        "Size": 10,
        "Product name": 34,
        "Local warehouse 101": 18,
        "Local warehouse 501": 18,
        "Central warehouse": 18,
        "Total units / style-colour": 24,
        "Selected warehouse availability": 26,
        "Available sizes / style-colour": 28,
        "Full size run": 14,
        "Top style": 12,
        "Material": 24,
        "ORDER": 12,
        "Discount %": 12,
        "RRP CZK": 13,
        "RRP EUR": 13,
        "Offer price CZK excl. VAT": 26,
        "Offer price EUR excl. VAT": 26,
        "Offer price CZK incl. VAT": 26,
        "Offer price EUR incl. VAT": 26,
        "EAN": 18,
        "Gender": 12,
        "Silhouette": 14,
        "Fit": 16,
        "End use": 16,
        "Season": 12,
        "C/O": 10,
        "Color group": 14,
        "Color name": 22,
        "Color code": 12,
        "Detail silhouette": 20,
        "Composition": 34,
        "Total availability (all warehouses)": 26,
        "Style code": 14,
    }
    for idx, column_name in enumerate(offer_df.columns, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = widths.get(column_name, 14)

    ws.row_dimensions[1].height = 22
    ws.row_dimensions[start_row].height = 30

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()

def unique_sorted(data: pd.DataFrame, col: str) -> list[str]:
    if col not in data.columns:
        return []
    values = [
        value
        for value in clean_text(data[col]).unique().tolist()
        if value and value.casefold() not in {"0", "nan", "none"}
    ]
    return sorted(values)
